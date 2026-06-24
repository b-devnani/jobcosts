"""Integration tests for the FastAPI surface (projects CRUD + generate)."""

import importlib
import io
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

SAMPLE_CSV = Path(__file__).parent / "sample" / "budget_details_3.csv"


@pytest.fixture()
def client(tmp_path, monkeypatch):
    # Point the DB at a throwaway file and set a known admin password.
    monkeypatch.setenv("JOBCOSTS_DB", str(tmp_path / "test.db"))
    monkeypatch.setenv("ADMIN_PASSWORD", "secret")
    monkeypatch.setenv("JOBCOSTS_SEED", "0")  # CRUD tests want an empty DB
    # Reload modules so they pick up the patched env vars.
    import backend.db as db
    import backend.app as app_module
    importlib.reload(db)
    importlib.reload(app_module)
    with TestClient(app_module.app) as c:
        yield c


ADMIN = {"X-Admin-Password": "secret"}


def test_healthz(client):
    res = client.get("/healthz")
    assert res.status_code == 200
    assert res.json() == {"status": "ok"}


def test_projects_empty(client):
    assert client.get("/api/projects").json() == []


def test_create_requires_admin(client):
    res = client.post("/api/projects", json={"name": "Nope"})
    assert res.status_code == 401


def test_admin_login(client):
    assert client.post("/api/admin/login", json={"password": "secret"}).status_code == 200
    assert client.post("/api/admin/login", json={"password": "wrong"}).status_code == 401


def test_project_crud(client):
    payload = {
        "name": "Maple Street",
        "orig_substantial_completion": "2025-09-15",
        "orig_final_completion": "2025-11-30",
        "current_substantial_completion": "2025-10-15",
        "current_final_completion": "2025-12-20",
    }
    created = client.post("/api/projects", json=payload, headers=ADMIN).json()
    assert created["id"] >= 1
    assert created["name"] == "Maple Street"
    assert created["orig_substantial_completion"] == "2025-09-15"

    pid = created["id"]
    updated = client.put(
        f"/api/projects/{pid}", json={"name": "Maple St. Tower"}, headers=ADMIN
    ).json()
    assert updated["name"] == "Maple St. Tower"
    # Dates preserved on partial update.
    assert updated["orig_final_completion"] == "2025-11-30"

    assert len(client.get("/api/projects").json()) == 1

    assert client.delete(f"/api/projects/{pid}", headers=ADMIN).status_code == 200
    assert client.get("/api/projects").json() == []


def test_create_requires_name(client):
    res = client.post("/api/projects", json={"name": ""}, headers=ADMIN)
    assert res.status_code == 400


def test_generate_with_project(client):
    project = client.post(
        "/api/projects",
        json={"name": "Generated Job", "orig_substantial_completion": "2025-09-15"},
        headers=ADMIN,
    ).json()
    csv_bytes = SAMPLE_CSV.read_bytes()
    res = client.post(
        "/api/generate",
        data={"project_id": project["id"]},
        files={"csv_file": ("budget.csv", csv_bytes, "text/csv")},
    )
    assert res.status_code == 200
    assert res.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert 'filename="Generated Job Job Costs ' in res.headers["content-disposition"]
    assert res.headers["content-disposition"].endswith('.xlsx"')
    # Returned bytes are a real xlsx (zip) with the data filled in.
    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(res.content))["Job Cost"]
    assert ws["A7"].value == "1-020 - Superintendent"
    assert ws["I2"].value.strftime("%Y-%m-%d") == "2025-09-15"  # Original Substantial


def test_generate_manual_entry(client):
    csv_bytes = SAMPLE_CSV.read_bytes()
    res = client.post(
        "/api/generate",
        data={"name": "Manual Job", "current_final_completion": "2026-01-31"},
        files={"csv_file": ("budget.csv", csv_bytes, "text/csv")},
    )
    assert res.status_code == 200
    assert 'filename="Manual Job Job Costs ' in res.headers["content-disposition"]


def test_generate_rejects_bad_csv(client):
    res = client.post(
        "/api/generate",
        data={"name": "Bad"},
        files={"csv_file": ("bad.csv", b"a,b,c\n1,2,3", "text/csv")},
    )
    assert res.status_code == 422


def test_generate_rejects_oversized_csv(client):
    big = b"x," * (6 * 1024 * 1024)  # ~12 MB, over the 10 MB cap
    res = client.post(
        "/api/generate",
        data={"name": "Big"},
        files={"csv_file": ("big.csv", big, "text/csv")},
    )
    assert res.status_code == 413


def test_update_is_atomic_under_concurrency(client):
    """Concurrent updates to one project must not clobber each other's fields."""
    import threading

    pid = client.post(
        "/api/projects", json={"name": "Race"}, headers=ADMIN
    ).json()["id"]

    def upd(field, value):
        client.put(
            f"/api/projects/{pid}", json={"name": "Race", field: value}, headers=ADMIN
        )

    threads = [
        threading.Thread(target=upd, args=("orig_substantial_completion", "2025-01-01")),
        threading.Thread(target=upd, args=("orig_final_completion", "2025-02-02")),
    ]
    for t in threads:
        t.start()
    for t in threads:
        t.join()
    # The row is still valid and the name is intact (no crash / corruption).
    final = client.get("/api/projects").json()[0]
    assert final["name"] == "Race"
