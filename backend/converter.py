"""Core logic: turn a Procore-style budget-detail CSV into a filled-in
Job Cost Projection workbook.

The transformation mirrors the manual Excel procedure documented in the
project brief (steps 6-9):

    6. Delete columns A, D and E of the exported CSV.
    7. Keep columns A (cost code) through I (job to date cost).
    8. Paste those columns, as values, into the template starting at row 8.
    9. Delete the unused template rows below the pasted data.

On top of the pasted data we also stamp the four project milestone dates and
name the workbook after the project, which is how the template's title cell
(an array formula reading the file name) shows the project name.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

TEMPLATE_PATH = Path(__file__).resolve().parent / "template" / "Job_Cost_Projection_Template.xlsx"

SHEET_NAME = "Job Cost"

# Row of the table header inside the template; data starts on the next row.
HEADER_ROW = 7
FIRST_DATA_ROW = 8
# Last data row the pristine template ships with (row 154 holds the totals).
LAST_TEMPLATE_DATA_ROW = 153
TOTALS_ROW = 154

# Header columns that carry currency values and must be stored as numbers so
# the workbook's own formulas keep working.
NUMERIC_OUTPUT_COLUMNS = ("C", "D", "E", "F", "G", "H", "I")

DATE_FORMAT = "mm-dd-yy"

# Visible width of the main table (columns A..M) used when applying styling.
TABLE_LAST_COL = 13  # column M
COMMITTED_COL = 7    # column G — Committed Costs (shown blue)
EAC_COL = 11         # column K — Estimated Cost at Completion (shown green)

# Brand + table palette (kept deliberately spare: the only fill is the grey
# table header; columns G/K carry colour, everything else is plain).
_BRAND_DARK = "26333F"      # BURLING charcoal — title / labels above the table
_LOGO_BLUE = "2E6DA4"       # BURLING logo box border + value underlines
_GREY_HEADER = "D9D9D9"     # table header fill
_COMMITTED_BLUE = "0070C0"  # Committed Costs values
_EAC_GREEN = "1E7B34"       # Estimated Cost at Completion values
_BORDER = "808080"          # table column / cell borders
_BLACK = "000000"
_THEME_FONT = "Century Gothic"  # geometric sans for the header area

# Indexes (0-based) of the CSV columns that survive "delete columns A, D, E"
# and form template columns A..I.  Order matters: it is the template order.
#
#   CSV layout (0-based):
#     0 Cost Code Tier 1      <- deleted (col A)
#     1 Cost Code Tier 2      -> template A  Cost Code/Description
#     2 Cost Type             -> template B  CAT
#     3 Budget Code           <- deleted (col D)
#     4 Budget Code Desc.     <- deleted (col E)
#     5 Original Budget       -> template C
#     6 Budget Modifications  -> template D
#     7 Approved COs          -> template E
#     8 Revised Budget        -> template F
#     9 Committed Costs       -> template G
#     10 Direct Cost          -> template H
#     11 Job to date Costs    -> template I
CSV_COLUMN_ORDER = [1, 2, 5, 6, 7, 8, 9, 10, 11]
# Which of the kept columns are text vs. numeric (by template column letter).
TEXT_COLUMNS = {"A", "B"}

# A handful of header names we verify so a non-Procore CSV that merely has 12+
# columns cannot map the wrong fields into the template and silently corrupt it.
EXPECTED_HEADERS = {
    1: "cost code tier 2",
    2: "cost type",
    5: "original budget amount",
    11: "job to date costs",
}


class ConversionError(ValueError):
    """Raised when the uploaded CSV does not look like a budget-detail export."""


@dataclass
class ProjectInfo:
    """The "remaining info" that does not come from the CSV.

    ``name`` becomes the workbook file name (and therefore the title cell).
    The four milestone dates land in their template cells; the last-pay-app
    figures fill the "Contract amount on last pay app and month" cells.
    """

    name: str
    orig_substantial_completion: date | None = None
    orig_final_completion: date | None = None
    current_substantial_completion: date | None = None
    current_final_completion: date | None = None
    contract_amount_last_pay_app: float | None = None
    month_last_pay_app: date | None = None


def _to_number(raw: str) -> float | str | None:
    """Parse a CSV money cell into a float, treating blanks as ``None``."""
    if raw is None:
        return None
    text = raw.strip().strip('"').replace(",", "").replace("$", "")
    if text == "" or text.lower() == "none":
        return None
    try:
        return float(text)
    except ValueError:
        # Non-numeric content in a money column -> leave as-is so nothing is lost.
        return raw.strip()


def _clean_text(raw: str | None) -> str:
    return (raw or "").strip()


def parse_budget_csv(content: str | bytes) -> list[list]:
    """Apply steps 6-7: drop columns A/D/E and keep A..I for each real row.

    Returns a list of 9-element rows in template column order (A..I).  Rows
    whose cost code is blank or the Procore ``None`` placeholder are skipped.
    """
    if isinstance(content, bytes):
        content = content.decode("utf-8-sig")

    reader = csv.reader(io.StringIO(content))
    rows = list(reader)
    if not rows:
        raise ConversionError("The CSV file is empty.")

    header = [c.strip().strip('"') for c in rows[0]]
    if len(header) < 12:
        raise ConversionError(
            "Unexpected CSV format: expected a Procore budget-detail export with "
            f"at least 12 columns, found {len(header)}."
        )
    mismatches = [
        f"column {idx} should be '{name}' but is "
        f"'{header[idx] if idx < len(header) else ''}'"
        for idx, name in EXPECTED_HEADERS.items()
        if idx >= len(header) or header[idx].strip().lower() != name
    ]
    if mismatches:
        raise ConversionError(
            "Unexpected CSV format (is this a Procore budget-detail export?): "
            + "; ".join(mismatches)
        )

    out: list[list] = []
    for raw_row in rows[1:]:
        if not any(cell.strip() for cell in raw_row):
            continue  # fully blank line
        if len(raw_row) <= max(CSV_COLUMN_ORDER):
            continue  # malformed / short line

        cost_code = _clean_text(raw_row[1])
        if cost_code == "" or cost_code.lower() == "none":
            continue  # Procore placeholder / subtotal row

        record = []
        for src_idx, col_letter in zip(CSV_COLUMN_ORDER, "ABCDEFGHI"):
            if col_letter in TEXT_COLUMNS:
                record.append(_clean_text(raw_row[src_idx]))
            else:
                record.append(_to_number(raw_row[src_idx]))
        out.append(record)

    if not out:
        raise ConversionError("No data rows were found in the CSV.")
    return out


def _coerce_number(value) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("$", "")
    if text == "":
        return None
    try:
        return float(text)
    except ValueError:
        raise ConversionError(f"Could not understand amount value: {value!r}")


def _coerce_date(value) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
        raise ConversionError(f"Could not understand date value: {value!r}")
    raise ConversionError(f"Unsupported date value: {value!r}")


def _set_date(ws: Worksheet, coord: str, value: date | None) -> None:
    if value is None:
        return
    cell = ws[coord]
    cell.value = value
    cell.number_format = DATE_FORMAT


def _set_number(ws: Worksheet, coord: str, value: float | None) -> None:
    if value is None:
        return
    ws[coord].value = value  # keeps the template's existing currency format


def _rewrite_formulas_for_totals(ws: Worksheet, data_rows: int) -> None:
    """After the surplus rows are deleted, repoint the formulas that referenced
    the original totals/summary rows so the workbook still recalculates."""
    totals_row = FIRST_DATA_ROW + data_rows  # new position of the totals row
    last_data = totals_row - 1

    # Totals row: SUM(col8:col<last_data>) for every numeric/derived column.
    for col in "CDEFGHIJKL":
        ws[f"{col}{totals_row}"] = f"=SUM({col}{FIRST_DATA_ROW}:{col}{last_data})"

    # Header contract-amount cells reference the totals row.
    ws["C3"] = f"=+C{totals_row}"   # Original Contract Amount  (Original Budget total)
    ws["C4"] = f"=+E{totals_row}"   # Approved PCCO's           (Approved COs total)
    ws["C5"] = f"=+F{totals_row}"   # Current Contract Amount   (Revised Budget total)

    # Summary block below the totals (PROJECT FEE ... Pending PCCO Fee) shifts up
    # by the number of deleted rows. Repoint its one internal formula.
    shift = LAST_TEMPLATE_DATA_ROW - last_data  # rows removed
    if shift:
        new_l159 = 159 - shift
        new_l160 = 160 - shift
        new_l161 = 161 - shift
        ws[f"L{new_l161}"] = f"=+L{new_l159}-L{new_l160}"


def _recolour(cell, rgb: str) -> None:
    """Change only a cell's font colour, keeping its other font attributes."""
    f = cell.font
    cell.font = Font(name=f.name, sz=f.sz, bold=f.bold, italic=f.italic, color=rgb)


def _style_workbook(ws: Worksheet, data_rows: int, title: str) -> None:
    """Apply the requested presentation on top of the filled-in template.

    Above the table: the BURLING logo + a themed title/summary. The table is
    left as-is apart from column borders, a grey header, blue Committed-Costs and
    green Estimated-Cost-at-Completion columns, and a bold totals row with a top
    border and double bottom border. Values, formulas and number formats are
    never touched.
    """
    totals_row = FIRST_DATA_ROW + data_rows
    last_col = TABLE_LAST_COL

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"

    # --- Logo (top-left) --- replaces the fragile array formula in A1 -------- #
    ws["A1"] = "BURLING"
    logo = ws["A1"]
    logo.font = Font(name="Times New Roman", size=20, bold=True, color=_BRAND_DARK)
    logo.alignment = Alignment(horizontal="center", vertical="center")
    blue = Side(style="medium", color=_LOGO_BLUE)
    logo.border = Border(left=blue, right=blue, top=blue, bottom=blue)
    ws.row_dimensions[1].height = 40

    # --- Title (themed font) ----------------------------------------------- #
    ws.merge_cells("C1:K1")
    t = ws["C1"]
    t.value = title or "Job Cost Projection"
    t.font = Font(name=_THEME_FONT, size=18, bold=True, color=_BRAND_DARK)
    t.alignment = Alignment(horizontal="center", vertical="center")

    # --- Summary block (rows 3-5): themed fonts, underlined values --------- #
    underline = Border(bottom=Side(style="thin", color=_LOGO_BLUE))
    for r in range(3, 6):
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                cell.font = Font(name=_THEME_FONT, size=11, bold=True, color=_BRAND_DARK)
    for coord in ("C3", "C4", "C5", "F3", "G3", "I3", "I4", "K3", "K4"):
        cell = ws[coord]
        cell.font = Font(name=_THEME_FONT, size=11, bold=True, color=_BRAND_DARK)
        cell.border = underline
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Table: borders + grey header (table content otherwise unchanged) --- #
    col_side = Side(style="thin", color=_BORDER)
    col_border = Border(left=col_side, right=col_side)
    header_box = Border(left=col_side, right=col_side, top=col_side, bottom=col_side)

    for c in range(1, last_col + 1):
        cell = ws.cell(row=HEADER_ROW, column=c)
        cell.fill = PatternFill("solid", fgColor=_GREY_HEADER)
        cell.font = Font(bold=True, sz=cell.font.sz or 11, name=cell.font.name, color=_BLACK)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_box

    for i in range(data_rows):
        r = FIRST_DATA_ROW + i
        for c in range(1, last_col + 1):
            ws.cell(row=r, column=c).border = col_border
        _recolour(ws.cell(row=r, column=COMMITTED_COL), _COMMITTED_BLUE)
        _recolour(ws.cell(row=r, column=EAC_COL), _EAC_GREEN)

    # --- Totals row: bold, top border + double bottom border --------------- #
    top = Side(style="thin", color=_BLACK)
    double = Side(style="double", color=_BLACK)
    if ws.cell(row=totals_row, column=1).value in (None, ""):
        ws.cell(row=totals_row, column=1).value = "TOTALS"
    for c in range(1, last_col + 1):
        cell = ws.cell(row=totals_row, column=c)
        cell.font = Font(bold=True, sz=cell.font.sz or 11, name=cell.font.name)
        cell.border = Border(left=col_side, right=col_side, top=top, bottom=double)
    _recolour(ws.cell(row=totals_row, column=COMMITTED_COL), _COMMITTED_BLUE)
    _recolour(ws.cell(row=totals_row, column=EAC_COL), _EAC_GREEN)

    # --- Summary block under the totals (bold labels, bordered values) ------ #
    for r in range(totals_row + 2, totals_row + 8):
        label_cell = ws.cell(row=r, column=11)   # column K
        value_cell = ws.cell(row=r, column=12)   # column L
        if label_cell.value is not None:
            label_cell.font = Font(bold=True, sz=label_cell.font.sz or 11, name=label_cell.font.name)
            value_cell.border = header_box


def _setup_print(ws: Worksheet, data_rows: int) -> None:
    """Landscape, narrow margins, fit all columns to one page wide, repeat the
    header rows, and a 'Page x of x' centre footer."""
    last_row = FIRST_DATA_ROW + data_rows + 7  # through the summary block

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(
        left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3, footer=0.3
    )
    ws.print_title_rows = "1:7"  # repeat title + summary + header on every page
    ws.print_area = f"A1:{get_column_letter(TABLE_LAST_COL)}{last_row}"
    ws.oddFooter.center.text = "Page &P of &N"
    ws.evenFooter.center.text = "Page &P of &N"


def build_workbook(csv_rows: Sequence[Sequence], project: ProjectInfo):
    """Return an openpyxl workbook: the template with data + dates filled in."""
    try:
        wb = load_workbook(TEMPLATE_PATH)
    except FileNotFoundError:
        raise ConversionError(f"Template workbook not found at {TEMPLATE_PATH}.")
    except Exception as exc:  # corrupt / unreadable workbook
        raise ConversionError(f"Template workbook could not be read: {exc}")
    if SHEET_NAME not in wb.sheetnames:
        raise ConversionError(
            f"Template is missing the required '{SHEET_NAME}' sheet."
        )
    ws = wb[SHEET_NAME]

    n = len(csv_rows)
    capacity = LAST_TEMPLATE_DATA_ROW - FIRST_DATA_ROW + 1  # 146 rows
    if n > capacity:
        raise ConversionError(
            f"The CSV has {n} rows but the template only has room for {capacity}."
        )

    # Step 8: paste columns A..I as values into the data rows.
    for i, record in enumerate(csv_rows):
        row = FIRST_DATA_ROW + i
        for col_letter, value in zip("ABCDEFGHI", record):
            ws[f"{col_letter}{row}"] = value

    # Step 9: delete the unused template rows below the pasted data, then fix
    # up the formulas that pointed at the (now moved) totals/summary rows.
    surplus_start = FIRST_DATA_ROW + n
    surplus_count = LAST_TEMPLATE_DATA_ROW - surplus_start + 1
    if surplus_count > 0:
        ws.delete_rows(surplus_start, surplus_count)
    _rewrite_formulas_for_totals(ws, n)

    # Stamp the four milestone dates from the project record.
    _set_date(ws, "I3", project.orig_substantial_completion)
    _set_date(ws, "I4", project.orig_final_completion)
    _set_date(ws, "K3", project.current_substantial_completion)
    _set_date(ws, "K4", project.current_final_completion)

    # "Contract amount on last pay app and month" header cells.
    _set_number(ws, "F3", project.contract_amount_last_pay_app)
    _set_date(ws, "G3", project.month_last_pay_app)

    # Make it presentable and print-ready.
    _style_workbook(ws, n, project.name)
    _setup_print(ws, n)

    return wb


def safe_filename(name: str, report_date: date | None = None, max_length: int = 150) -> str:
    """Build the download name: ``<Project Name> Job Costs MMDDYY.xlsx``.

    The project name is sanitised (and trimmed if very long); the date defaults
    to today.
    """
    report_date = report_date or date.today()
    keep = "".join(c if c.isalnum() or c in " -_." else "_" for c in name).strip()
    base = keep or "Job Cost Projection"
    suffix = f" Job Costs {report_date.strftime('%m%d%y')}.xlsx"
    if len(base) + len(suffix) > max_length:
        base = base[: max_length - len(suffix)].strip()
    return base + suffix


def convert_csv_to_workbook_bytes(
    csv_content: str | bytes,
    name: str,
    orig_substantial_completion=None,
    orig_final_completion=None,
    current_substantial_completion=None,
    current_final_completion=None,
    contract_amount_last_pay_app=None,
    month_last_pay_app=None,
    report_date=None,
) -> tuple[bytes, str]:
    """High-level entry point used by the web layer.

    Returns ``(xlsx_bytes, download_filename)``.
    """
    project = ProjectInfo(
        name=name,
        orig_substantial_completion=_coerce_date(orig_substantial_completion),
        orig_final_completion=_coerce_date(orig_final_completion),
        current_substantial_completion=_coerce_date(current_substantial_completion),
        current_final_completion=_coerce_date(current_final_completion),
        contract_amount_last_pay_app=_coerce_number(contract_amount_last_pay_app),
        month_last_pay_app=_coerce_date(month_last_pay_app),
    )
    rows = parse_budget_csv(csv_content)
    wb = build_workbook(rows, project)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), safe_filename(name, report_date)
