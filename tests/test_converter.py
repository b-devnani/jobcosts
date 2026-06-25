"""Tests for the CSV -> Job Cost Projection workbook conversion."""

import io
import re
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from backend import converter

SAMPLE_CSV = Path(__file__).parent / "sample" / "budget_details_3.csv"


@pytest.fixture(scope="module")
def csv_text():
    return SAMPLE_CSV.read_text()


@pytest.fixture(scope="module")
def parsed(csv_text):
    return converter.parse_budget_csv(csv_text)


# --------------------------------------------------------------------------- #
# Parsing (steps 6-7)
# --------------------------------------------------------------------------- #
def test_parse_drops_placeholder_rows(parsed):
    # The sample has 64 data lines; one is the Procore "None" placeholder.
    assert len(parsed) == 63
    assert all(str(r[0]).strip().lower() not in ("", "none") for r in parsed)


def test_parse_keeps_nine_columns(parsed):
    assert all(len(r) == 9 for r in parsed)


def test_parse_column_mapping(parsed):
    first = parsed[0]
    # A=cost code, B=CAT, C=original budget ... I=job to date
    assert first[0] == "1-020 - Superintendent"
    assert first[1] == "L - Labor"
    assert first[2] == 134000.0          # Original Budget Amount
    assert first[3] == -30500.0          # Budget Modifications
    assert first[8] == 39358.6           # Job to date Costs


def test_numeric_columns_are_numbers(parsed):
    for row in parsed:
        for value in row[2:]:
            assert value is None or isinstance(value, (int, float))


def test_bad_csv_raises():
    with pytest.raises(converter.ConversionError):
        converter.parse_budget_csv("not,a,budget,export\n1,2,3,4")


def test_empty_csv_raises():
    with pytest.raises(converter.ConversionError):
        converter.parse_budget_csv("")


def test_spoofed_csv_with_wrong_columns_is_rejected():
    # 12 columns, column 1 is "Cost Code Tier 2", but the other key columns are
    # wrong -> must be rejected instead of silently mapping the wrong fields.
    header = (
        "Tier1,Cost Code Tier 2,WRONG Type,Budget Code,Desc,WRONG Budget,"
        "Mods,COs,Revised,Committed,Direct,WRONG JTD\n"
    )
    row = "1,1-020 - X,L,code,d,100,0,0,100,0,50,50\n"
    with pytest.raises(converter.ConversionError) as exc:
        converter.parse_budget_csv(header + row)
    assert "column 2" in str(exc.value) or "column 5" in str(exc.value)


# --------------------------------------------------------------------------- #
# Workbook building (steps 8-9 + milestones)
# --------------------------------------------------------------------------- #
@pytest.fixture(scope="module")
def workbook_bytes(csv_text):
    data, fname = converter.convert_csv_to_workbook_bytes(
        csv_text,
        name="Maple Street",
        orig_substantial_completion="2025-09-15",
        orig_final_completion="2025-11-30",
        current_substantial_completion="2025-10-15",
        current_final_completion="2025-12-20",
        report_date=date(2026, 6, 24),
    )
    return data, fname


def _ws(workbook_bytes):
    data, _ = workbook_bytes
    return load_workbook(io.BytesIO(data))["Job Cost"]


def test_filename_format(workbook_bytes):
    _, fname = workbook_bytes
    assert fname == "Maple Street Job Costs 062426.xlsx"


def test_data_pasted_at_row_7(workbook_bytes):
    ws = _ws(workbook_bytes)
    assert ws["A7"].value == "1-020 - Superintendent"
    assert ws["B7"].value == "L - Labor"
    assert ws["C7"].value == 134000.0


def test_surplus_rows_deleted(workbook_bytes, parsed):
    ws = _ws(workbook_bytes)
    totals_row = 7 + len(parsed)  # 70
    # The row right after the data is the totals row, not an empty data row.
    assert ws[f"C{totals_row}"].value == f"=SUM(C7:C{totals_row - 1})"
    # No data rows survive beyond the totals row except the summary block.
    assert ws[f"A{totals_row - 1}"].value == parsed[-1][0]


def test_totals_and_header_formulas_repointed(workbook_bytes, parsed):
    ws = _ws(workbook_bytes)
    totals_row = 7 + len(parsed)
    assert ws["C2"].value == f"=+C{totals_row}"
    assert ws["C3"].value == f"=+E{totals_row}"
    assert ws["C4"].value == f"=+F{totals_row}"


def test_etc_is_value_not_formula(workbook_bytes, parsed):
    """Estimated Cost to Complete (col J) must be a computed value, not a
    formula. K (=I+J) and L (=F-K) remain formulas that read it."""
    ws = _ws(workbook_bytes)
    last = 6 + len(parsed)
    rec = parsed[-1]
    revised, committed, jtd = rec[5] or 0, rec[6] or 0, rec[8] or 0
    expected = round(max(revised, committed) - jtd, 2)
    assert ws[f"J{last}"].value == expected
    assert not (isinstance(ws[f"J{last}"].value, str) and ws[f"J{last}"].value.startswith("="))
    assert ws[f"K{last}"].value == f"=+I{last}+J{last}"
    assert ws[f"L{last}"].value == f"=+F{last}-K{last}"


def test_milestone_dates_written(workbook_bytes):
    ws = _ws(workbook_bytes)
    assert ws["I2"].value.date() == date(2025, 9, 15)   # Original Substantial
    assert ws["I3"].value.date() == date(2025, 11, 30)  # Original Final
    assert ws["K2"].value.date() == date(2025, 10, 15)  # Current Substantial
    assert ws["K3"].value.date() == date(2025, 12, 20)  # Current Final


def test_no_dangling_formula_references(workbook_bytes):
    ws = _ws(workbook_bytes)
    max_row = ws.max_row
    ref = re.compile(r"\$?[A-Z]{1,2}\$?(\d+)")
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                for m in ref.finditer(cell.value):
                    assert int(m.group(1)) <= max_row, (cell.coordinate, cell.value)


def test_no_milestones_leaves_cells_blank(csv_text):
    data, _ = converter.convert_csv_to_workbook_bytes(csv_text, name="No Dates")
    ws = load_workbook(io.BytesIO(data))["Job Cost"]
    assert ws["I2"].value is None
    assert ws["K3"].value is None
    assert ws["F2"].value is None          # contract amount on last pay app
    assert ws["G2"].value is None          # month of last pay app


def test_last_pay_app_fields_written(csv_text):
    data, _ = converter.convert_csv_to_workbook_bytes(
        csv_text,
        name="Pay App",
        contract_amount_last_pay_app="1,234,567.89",
        month_last_pay_app="2026-05-31",
    )
    ws = load_workbook(io.BytesIO(data))["Job Cost"]
    assert ws["F2"].value == 1234567.89
    assert ws["G2"].value.date() == date(2026, 5, 31)
    # Month keeps the template's own "mmm 'yy" format, not the date format.
    assert "mmm" in ws["G2"].number_format


def test_safe_filename():
    d = date(2026, 6, 24)
    assert converter.safe_filename("A/B:C*?", d) == "A_B_C__ Job Costs 062426.xlsx"
    assert converter.safe_filename("   ", d) == "Job Cost Projection Job Costs 062426.xlsx"


def test_safe_filename_defaults_to_today():
    out = converter.safe_filename("Riverside")
    assert out.startswith("Riverside Job Costs ")
    assert out.endswith(".xlsx")


def test_safe_filename_truncates_long_names():
    out = converter.safe_filename("X" * 500, date(2026, 6, 24))
    assert len(out) <= 150
    assert out.endswith(" Job Costs 062426.xlsx")


def test_no_array_formula_regression(workbook_bytes):
    """A1's array formula caused Excel's 'we found a problem' repair prompt; the
    template must contain no array formula at all."""
    import zipfile
    data, _ = workbook_bytes
    with zipfile.ZipFile(io.BytesIO(data)) as z:
        sheet = next(n for n in z.namelist() if n.endswith("sheet1.xml"))
        xml = z.read(sheet).decode("utf-8", "replace")
    assert 't="array"' not in xml


def test_title_is_download_filename(workbook_bytes):
    ws = _ws(workbook_bytes)
    assert ws["A1"].value == "Maple Street Job Costs 062426"


def test_logo_embedded(workbook_bytes):
    import zipfile
    data, _ = workbook_bytes
    with zipfile.ZipFile(io.BytesIO(data)) as z:
        assert any("media/image" in n for n in z.namelist())


def test_workbook_is_styled(workbook_bytes, parsed):
    ws = _ws(workbook_bytes)
    totals_row = 7 + len(parsed)
    # Grey header row (the only fill) at row 6.
    assert ws["A6"].fill.fgColor.rgb.endswith("D9D9D9")
    assert ws["A6"].font.bold
    assert ws["A7"].fill.patternType is None  # data rows have no fill
    # Committed Costs column blue, Estimated Cost at Completion column green.
    assert ws["G7"].font.color.rgb.endswith("0070C0")
    assert ws["K7"].font.color.rgb.endswith("1E7B34")
    # Column borders present on data cells.
    assert ws["C7"].border.left.style == "thin"
    # Totals row: labelled, bold, top + double bottom borders.
    assert ws.cell(row=totals_row, column=1).value == "TOTALS"
    assert ws.cell(row=totals_row, column=3).font.bold
    assert ws.cell(row=totals_row, column=3).border.top.style == "thin"
    assert ws.cell(row=totals_row, column=3).border.bottom.style == "double"


def test_print_settings(workbook_bytes):
    ws = _ws(workbook_bytes)
    assert ws.page_setup.orientation == "landscape"
    assert ws.page_setup.fitToWidth == 1
    assert ws.print_title_rows in ("1:6", "$1:$6")
    assert ws.oddFooter.center.text == "Page &P of &N"
    assert "$A$1:$M$" in ws.print_area
